"""Mode chauffeur (terrain) : constat en DEUX temps, cases par partie, 2 signatures, PDF.

L'état des lieux se fait en deux moments sur LE MÊME document :
- « Début de location » (au départ du module) : colonne C, signature côté GAUCHE (col A).
- « Fin de location »   (au retour du module)  : colonne F, signature côté DROITE (col E).

Pour chaque PARTIE du module (murs, porte, coffret électrique, sol…) — lues directement dans le
formulaire, donc adaptées à chaque type de bloc — le chauffeur choisit l'état : Bon / Sale / Cassé
(+ note). Le texte est reporté dans la bonne colonne de l'Excel (logo et perspectives préservés via
patch_xlsx). Chaque phase, une fois signée, est FIGÉE (preuve). L'état de saisie vit dans
constat.json ; le dossier « constat - <document> » reçoit photos, signatures et PDF.
"""

from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app import patch_xlsx

PHOTOS_MAX = 20
_POINTILLES = {".", "…", " ", " "}
_COL_DEBUT, _COL_FIN = 3, 6  # colonnes C (début de loc) et F (fin de loc)
PHASES = ("debut", "fin")

# Libellés d'état reportés dans l'Excel (le choix « clé » de l'UI → texte du document).
ETATS = {"bon": "Bon état", "sale": "Sale", "casse": "Cassé", "reparer": "À réparer"}


def _est_pointille(v) -> bool:
    """True si la cellule ne contient qu'un gabarit de pointillés (case à remplir)."""
    return isinstance(v, str) and len(v.strip()) > 3 and set(v.strip()) <= _POINTILLES


def dossier_constat(job_dir: Path, nom_fichier: str) -> Path:
    """Dossier des pièces du constat (photos, signatures, PDF), accolé au document."""
    d = job_dir / f"constat - {Path(nom_fichier).stem}"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _chemin_json(dossier: Path) -> Path:
    return dossier / "constat.json"


def _feuille(wb, feuille):
    return wb[feuille] if feuille and feuille in wb.sheetnames else wb[wb.sheetnames[0]]


def analyser_parties(chemin: Path, feuille: str | None) -> list[dict]:
    """PARTIES à contrôler, lues dans le formulaire et REGROUPÉES : une entrée par partie,
    même si elle couvre plusieurs lignes du modèle (ex. « Coffret électrique » sur 2 lignes).

    S'adapte automatiquement au type de bloc (chaque état des lieux liste ses propres parties)."""
    wb = load_workbook(chemin, data_only=True)
    ws = _feuille(wb, feuille)
    parties: list[dict] = []
    courant: dict | None = None
    for r in range(1, ws.max_row + 1):
        c = ws.cell(row=r, column=_COL_DEBUT).value
        f = ws.cell(row=r, column=_COL_FIN).value
        if not (_est_pointille(c) or _est_pointille(f)):
            continue
        a = ws.cell(row=r, column=1).value
        libelle = " ".join(str(a).split()) if isinstance(a, str) and str(a).strip() else ""
        if libelle:  # nouvelle partie
            courant = {"cle": f"r{r}", "libelle": libelle, "lignes": [r]}
            parties.append(courant)
        elif courant is not None:  # suite de la partie précédente (même case, autre ligne)
            courant["lignes"].append(r)
        else:  # ligne à remplir sans libellé au tout début du tableau
            courant = {"cle": f"r{r}", "libelle": f"Partie (ligne {r})", "lignes": [r]}
            parties.append(courant)
    wb.close()
    return parties


def _texte_etat(etat: str, note: str) -> str:
    """« casse » + « vitre fissurée » -> « Cassé — vitre fissurée » (texte écrit dans l'Excel)."""
    libelle = ETATS.get((etat or "").strip(), "")
    note = (note or "").strip()
    if not libelle:
        return note  # état non choisi mais note libre : on écrit la note seule
    return f"{libelle} — {note}" if note else libelle


def _phase_bloc(data: dict, phase: str) -> dict:
    """Bloc de signature d'une phase (vide si absent)."""
    return data.get(phase) if isinstance(data.get(phase), dict) else {}


def charger_constat(document: Path, feuille: str | None, dossier: Path) -> dict:
    """État complet : parties (avec saisie début/fin), photos, et l'état des 2 signatures."""
    js = _chemin_json(dossier)
    data = json.loads(js.read_text(encoding="utf-8")) if js.exists() else {}
    parties = data.get("parties")
    if not parties:
        parties = analyser_parties(document, feuille)
    for p in parties:  # garantit la présence des sous-blocs de saisie
        p.setdefault("debut", {"etat": "", "note": ""})
        p.setdefault("fin", {"etat": "", "note": ""})
    out: dict = {"feuille": feuille, "parties": parties,
                 "photos": sorted(p.name for p in dossier.glob("photo-*.*")),
                 "pdf": "constat.pdf" if (dossier / "constat.pdf").exists() else None}
    for ph in PHASES:
        s = _phase_bloc(data, ph)
        out[ph] = {"signe": bool(s.get("signe_le")), "signataire": s.get("signataire", ""),
                   "fonction": s.get("fonction", ""), "signe_le": s.get("signe_le", "")}
    return out


class ConstatSigne(Exception):
    """Levée quand on modifie une phase déjà signée (preuve figée)."""


def enregistrer_constat(document: Path, feuille: str | None, phase: str,
                        saisies: dict, dossier: Path) -> None:
    """Écrit l'état de chaque partie pour UNE phase dans la bonne colonne (C=début, F=fin).

    `saisies` = { cle_partie: {"etat": "bon|sale|casse|reparer", "note": "…"} }.
    Refuse si la phase est déjà signée (empreinte de preuve à préserver)."""
    if phase not in PHASES:
        raise ValueError("Phase invalide (attendu « debut » ou « fin »).")
    js = _chemin_json(dossier)
    data = json.loads(js.read_text(encoding="utf-8")) if js.exists() else {}
    if _phase_bloc(data, phase).get("signe_le"):
        raise ConstatSigne(
            f"La partie « {'début' if phase == 'debut' else 'fin'} de location » est déjà signée : "
            "elle ne peut plus être modifiée. Faites re-signer le client si une correction s'impose."
        )
    parties = data.get("parties") or analyser_parties(document, feuille)
    for p in parties:
        p.setdefault("debut", {"etat": "", "note": ""})
        p.setdefault("fin", {"etat": "", "note": ""})
    par_cle = {p["cle"]: p for p in parties}
    colonne = "C" if phase == "debut" else "F"
    valeurs: dict[str, object] = {}
    for cle, v in (saisies or {}).items():
        p = par_cle.get(cle)
        if not p:
            continue
        etat = (v.get("etat") or "").strip()
        note = (v.get("note") or "").strip()
        p[phase] = {"etat": etat, "note": note}
        texte = _texte_etat(etat, note)
        if texte:
            for r in p["lignes"]:
                valeurs[f"{colonne}{r}"] = texte
    if valeurs:
        patch_xlsx.ecrire_cellules(document, document, feuille, valeurs)
    data["feuille"] = feuille
    data["parties"] = parties
    data["maj"] = datetime.now().isoformat(timespec="seconds")
    js.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")


def ajouter_photo(dossier: Path, contenu: bytes) -> str:
    """Enregistre une photo (JPEG/PNG vérifiés par signature de fichier)."""
    if contenu.startswith(b"\xff\xd8\xff"):
        ext = "jpg"
    elif contenu.startswith(b"\x89PNG\r\n\x1a\n"):
        ext = "png"
    else:
        raise ValueError("Format accepté : JPEG ou PNG.")
    existantes = list(dossier.glob("photo-*.*"))
    if len(existantes) >= PHOTOS_MAX:
        raise ValueError(f"Maximum {PHOTOS_MAX} photos par constat.")
    numeros = [int(m.group(1)) for p in existantes if (m := re.match(r"photo-(\d+)", p.stem))]
    nom = f"photo-{(max(numeros) + 1) if numeros else 1:02d}.{ext}"
    (dossier / nom).write_bytes(contenu)
    return nom


def _ancres_signature(document: Path, feuille: str | None) -> dict:
    """Repère les DEUX zones « Date, Nom et Signature » : gauche (col A ≈ départ/début) et
    droite (col E ≈ retour/fin). Renvoie {"debut": ref_gauche, "fin": ref_droite}."""
    wb = load_workbook(document, data_only=True)
    ws = _feuille(wb, feuille)
    reperes: list[tuple[int, int, str]] = []
    for row in ws.iter_rows(max_col=8):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "Signature" in v and "mention" not in v:
                reperes.append((cell.column, cell.row, cell.column_letter))
    wb.close()
    if not reperes:
        return {}
    reperes.sort()  # tri par colonne puis ligne
    gauche, droite = reperes[0], reperes[-1]
    return {"debut": f"{gauche[2]}{gauche[1] + 1}", "fin": f"{droite[2]}{droite[1] + 1}"}


def enregistrer_signature(
    dossier: Path, png: bytes, signataire: str, document: Path | None = None,
    feuille: str | None = None, fonction: str = "", accord: bool = False, phase: str = "debut",
) -> str | None:
    """Enregistre la signature d'UNE phase + son dossier de preuve (identité, fonction, horodatage,
    empreinte SHA-256 du document au moment de la signature). Insère l'image dans la bonne zone du
    document (gauche pour le départ, droite pour le retour). Renvoie l'empreinte (None sans document)."""
    if phase not in PHASES:
        raise ValueError("Phase invalide.")
    if not png.startswith(b"\x89PNG\r\n\x1a\n"):
        raise ValueError("Signature invalide.")
    (dossier / f"signature-{phase}.png").write_bytes(png)
    empreinte: str | None = None
    if document is not None and document.exists():
        try:
            ancre = _ancres_signature(document, feuille).get(phase)
            if ancre:
                patch_xlsx.inserer_image(document, feuille, ancre, png, cle=phase)
        except Exception:  # noqa: BLE001 — l'insertion Excel ne doit pas bloquer la signature
            pass
        try:
            empreinte = hashlib.sha256(document.read_bytes()).hexdigest()
        except OSError:
            empreinte = None
    js = _chemin_json(dossier)
    data = json.loads(js.read_text(encoding="utf-8")) if js.exists() else {}
    now = datetime.now()
    bloc = {
        "signataire": (signataire or "").strip(),
        "fonction": (fonction or "").strip(),
        "accord": bool(accord),
        "signe_le": now.strftime("%d/%m/%Y à %H:%M"),
        "signe_le_iso": now.isoformat(timespec="seconds"),
    }
    if empreinte:
        bloc["empreinte_sha256"] = empreinte
    data[phase] = bloc
    js.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")
    return empreinte


def generer_pdf(dossier: Path, titre: str, sous_titre: str, societe: str = "") -> Path:
    """PDF de constat : relevés par partie (état début / état fin), photos, et les 2 signatures."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfgen import canvas as rl_canvas

    js = _chemin_json(dossier)
    data = json.loads(js.read_text(encoding="utf-8")) if js.exists() else {"parties": []}
    cible = dossier / "constat.pdf"
    c = rl_canvas.Canvas(str(cible), pagesize=A4)
    largeur, hauteur = A4
    marge = 18 * mm

    def entete_page() -> float:
        c.setFont("Helvetica-Bold", 15)
        c.drawString(marge, hauteur - marge, f"Constat d'état des lieux — {societe or 'GB Location'}")
        c.setFont("Helvetica", 10)
        c.drawString(marge, hauteur - marge - 14, titre)
        c.drawString(marge, hauteur - marge - 26, sous_titre)
        c.setFillColor(colors.grey)
        c.drawRightString(largeur - marge, hauteur - marge, datetime.now().strftime("%d/%m/%Y %H:%M"))
        c.setFillColor(colors.black)
        c.line(marge, hauteur - marge - 32, largeur - marge, hauteur - marge - 32)
        return hauteur - marge - 46

    y = entete_page()
    c.setFont("Helvetica-Bold", 9)
    c.drawString(marge, y, "Partie du module")
    c.drawString(marge + 78 * mm, y, "Début de location")
    c.drawString(marge + 128 * mm, y, "Fin de location")
    y -= 12
    c.setFont("Helvetica", 9)
    for p in data.get("parties", []):
        if y < 30 * mm:
            c.showPage()
            y = entete_page()
            c.setFont("Helvetica", 9)
        deb = _texte_etat((p.get("debut") or {}).get("etat", ""), (p.get("debut") or {}).get("note", ""))
        fin = _texte_etat((p.get("fin") or {}).get("etat", ""), (p.get("fin") or {}).get("note", ""))
        c.drawString(marge, y, str(p.get("libelle", ""))[:52])
        c.drawString(marge + 78 * mm, y, deb[:34])
        c.drawString(marge + 128 * mm, y, fin[:34])
        y -= 11

    photos = sorted(dossier.glob("photo-*.*"))
    if photos:
        c.showPage()
        y = entete_page()
        c.setFont("Helvetica-Bold", 11)
        c.drawString(marge, y, f"Photos ({len(photos)})")
        larg_img = (largeur - 2 * marge - 8 * mm) / 2
        haut_img = 62 * mm
        x = marge
        col = 0
        y -= 8 + haut_img
        for p in photos:
            try:
                c.drawImage(ImageReader(str(p)), x, y, width=larg_img, height=haut_img,
                            preserveAspectRatio=True, anchor="sw")
                c.setFont("Helvetica", 7)
                c.setFillColor(colors.grey)
                c.drawString(x, y - 8, p.name)
                c.setFillColor(colors.black)
            except Exception:  # noqa: BLE001
                continue
            col += 1
            if col % 2 == 0:
                x = marge
                y -= haut_img + 14
                if y < 30 * mm:
                    c.showPage()
                    y = entete_page() - haut_img
            else:
                x = marge + larg_img + 8 * mm

    # --- Les deux signatures (départ à gauche, retour à droite) ---
    signatures = [("debut", "Début de location (départ)", marge),
                  ("fin", "Fin de location (retour)", marge + (largeur - 2 * marge) / 2)]
    if any(_phase_bloc(data, ph).get("signe_le") for ph, _, _ in signatures):
        c.showPage()
        y = entete_page()
        c.setFont("Helvetica-Bold", 12)
        c.drawString(marge, y, "Signatures du client")
        for phase, lbl, x0 in signatures:
            b = _phase_bloc(data, phase)
            c.setFont("Helvetica-Bold", 10)
            c.drawString(x0, y - 22, lbl)
            c.setFont("Helvetica", 9)
            if b.get("signe_le"):
                qui = b.get("signataire") or ""
                fct = b.get("fonction") or ""
                c.drawString(x0, y - 36, (f"{qui} ({fct})" if fct else qui)[:44])
                mention = "« bon pour accord »" if b.get("accord") else "signature recueillie"
                c.drawString(x0, y - 48, f"{mention} · {b.get('signe_le', '')}")
                sig = dossier / f"signature-{phase}.png"
                if sig.exists():
                    try:
                        c.drawImage(ImageReader(str(sig)), x0, y - 110, width=64 * mm, height=26 * mm,
                                    preserveAspectRatio=True, anchor="sw", mask="auto")
                    except Exception:  # noqa: BLE001
                        pass
                if b.get("empreinte_sha256"):
                    c.setFont("Helvetica", 5.5)
                    c.setFillColor(colors.grey)
                    c.drawString(x0, y - 118, f"SHA-256 : {b['empreinte_sha256'][:48]}…")
                    c.setFillColor(colors.black)
            else:
                c.setFillColor(colors.grey)
                c.drawString(x0, y - 36, "— non signé —")
                c.setFillColor(colors.black)
    c.save()
    return cible
