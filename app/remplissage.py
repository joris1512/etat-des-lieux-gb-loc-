"""Remplissage des modèles Excel — en-tête + fonction + mobilier.

On écrit UNIQUEMENT les cellules prévues par config/cellules.yaml via `app.patch_xlsx`, qui
modifie le XML de la feuille ciblée **en préservant tout le reste du classeur** : logo GB,
dessins en perspective, mise en forme, autres onglets. (openpyxl, lui, perd images et dessins
à l'enregistrement — d'où ce moteur dédié.) On ne touche jamais aux champs d'état réel,
réserves ou signatures.
"""

from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path

import yaml
from openpyxl import load_workbook

from app import patch_xlsx, regles_modeles
from app.config import CONFIG_CELLULES
from app.correspondance import normaliser
from app.models import EnteteDevis, EtatDesLieux


@lru_cache
def _config_cellules() -> dict:
    return yaml.safe_load(CONFIG_CELLULES.read_text(encoding="utf-8"))


def _cfg_modele(modele: str) -> dict:
    """Config résolue pour un modèle : défaut + surcharges éventuelles."""
    cfg = _config_cellules()
    defaut = cfg.get("defaut", {}) or {}
    surcharge = (cfg.get("modeles", {}) or {}).get(modele, {}) or {}
    resolu = {
        "feuille": surcharge.get("feuille", defaut.get("feuille")),
        "fonction": surcharge.get("fonction", defaut.get("fonction")),
    }
    # En-tête, format d'en-tête et mobilier : une surcharge explicite (même vide) REMPLACE le
    # défaut (les vrais modèles ont leurs propres cellules, à ne pas fusionner avec le défaut).
    resolu["entete"] = (
        surcharge["entete"] if "entete" in surcharge else (defaut.get("entete") or {})
    )
    resolu["entete_format"] = (
        surcharge["entete_format"]
        if "entete_format" in surcharge
        else (defaut.get("entete_format") or {})
    )
    resolu["fonctions_cellules"] = (
        surcharge.get("fonctions_cellules") or defaut.get("fonctions_cellules") or {}
    )
    resolu["mobilier"] = (
        surcharge["mobilier"] if "mobilier" in surcharge else (defaut.get("mobilier") or {})
    )
    # Zone d'inventaire mobilier des VRAIS modèles : liste de cellules qui reçoivent la liste
    # COMPLÈTE du devis, telle quelle (« MICRO ONDES : 2 ») — rien n'est filtré ni perdu.
    resolu["mobilier_zone"] = (
        surcharge.get("mobilier_zone") or defaut.get("mobilier_zone") or []
    )
    # R2 : emplacements WC (« V : … ») à garder/effacer selon le nombre de WC de l'état.
    resolu["wc_emplacements"] = (
        surcharge.get("wc_emplacements") or defaut.get("wc_emplacements") or []
    )
    # Kitchenette (réfectoire) : cases évier/CE/micro/frigo à remplir avec la quantité du devis (0 si absent).
    resolu["kitchenette"] = surcharge.get("kitchenette") or defaut.get("kitchenette") or []
    # Lignes dont on garantit une hauteur mini (lisibilité impression, ex. doses / mise en eau).
    resolu["hauteur_lignes"] = surcharge.get("hauteur_lignes") or defaut.get("hauteur_lignes") or []
    # Cuve : cellules taille (3m/6m) et raccords (branchement).
    resolu["cuve"] = surcharge.get("cuve") or {}
    # Titre du bloc (règle Joris) : cellule + gabarit pour réécrire le titre avec le contenu réel
    # du bloc lu au devis (ex. polysani : A6 « Sanitaire : BLOC 5WC / 3UR (POLYSANI) »).
    resolu["titre_contenu"] = surcharge.get("titre_contenu") or defaut.get("titre_contenu") or {}
    return resolu


@lru_cache
def _libelles_mobilier() -> tuple[tuple[str, str], ...]:
    """Table (motif normalisé, libellé court) triée du motif le plus long au plus court."""
    entrees = []
    for e in _config_cellules().get("mobilier_libelles") or []:
        motif = normaliser(str(e.get("motif") or ""))
        libelle = str(e.get("libelle") or "").strip()
        if motif and libelle:
            entrees.append((motif, libelle))
    entrees.sort(key=lambda x: -len(x[0]))
    return tuple(entrees)


def _libelle_court(designation: str) -> str:
    """Libellé « norme GB » d'un article (ex. « Tables ») ; désignation d'origine sinon."""
    cible = normaliser(designation)
    for motif, libelle in _libelles_mobilier():
        if motif in cible:
            return libelle
    return designation.strip()


def _lignes_inventaire(mobilier) -> list[str]:
    """Liste « Libellé : quantité » : articles regroupés par libellé court, quantités sommées,
    ordre d'apparition du devis préservé."""
    totaux: dict[str, int] = {}
    for item in mobilier:
        cle = _libelle_court(item.designation)
        totaux[cle] = totaux.get(cle, 0) + item.quantite
    return [f"{libelle} : {qte}" for libelle, qte in totaux.items()]


def _cellule_mobilier(designation: str, table: dict[str, str]) -> str | None:
    """Trouve la cellule pour une désignation via mot-clé (le plus long qui correspond gagne)."""
    cible = normaliser(designation)
    candidats = sorted(table.items(), key=lambda kv: len(normaliser(kv[0])), reverse=True)
    for mot, cellule in candidats:
        if normaliser(mot) in cible:
            return cellule
    return None


def resumer_contenu_bloc(texte: str) -> str:
    """Décrit le contenu réel d'un bloc sanitaire lu sur la ligne du devis, au format des modèles
    GB (« 5WC / 3UR / 4 PTS EAU »). Sert à réécrire le titre de l'état des lieux (règle Joris).
    Renvoie "" si rien n'est reconnu (dans ce cas on garde le titre figé du modèle)."""
    t = normaliser(texte or "")

    def _n(motif: str) -> int:
        m = re.search(motif, t)
        return int(m.group(1)) if m else 0

    parties: list[str] = []
    if wc := _n(r"(\d+)\s*WC\b"):
        parties.append(f"{wc}WC")
    if ur := _n(r"(\d+)\s*URIN"):
        parties.append(f"{ur}UR")
    if douches := _n(r"(\d+)\s*DOUCHE"):
        parties.append(f"{douches} DOUCHES")
    if lavabos := _n(r"(\d+)\s*(?:LAVABO|LAVE\s*MAIN)"):
        parties.append(f"{lavabos} LAVABOS")
    if pts_eau := _n(r"(\d+)\s*(?:PTS?|POINTS?)\s*(?:D\s*)?EAU"):
        parties.append(f"{pts_eau} PTS EAU")
    return " / ".join(parties)


def remplir_etat(
    modele_path: Path,
    sortie_path: Path,
    entete: EnteteDevis,
    etat: EtatDesLieux,
) -> list[str]:
    """Remplit un modèle et l'enregistre. Renvoie la liste des désignations non mappées."""
    cfg = _cfg_modele(etat.modele)
    a_ecrire: dict[str, object] = {}

    # --- Feuille + cellules d'en-tête (cas DOUCHES : onglet 4D/6D + en-tête selon nb_douches, R1) ---
    feuille = cfg.get("feuille")
    entete_cellules = dict(cfg.get("entete") or {})
    if etat.nb_douches:
        if etat.nb_douches < 5:  # 4 douches -> Petit Sanitaire Douche (en-tête E4/E6)
            feuille, entete_cellules = "Petit Sanitaire Douche", {"client": "E4", "titre_chantier": "E6"}
        else:  # 6 douches -> Grand Sanitaire Douche (en-tête E5/E7)
            feuille, entete_cellules = "Grand Sanitaire Douche", {"client": "E5", "titre_chantier": "E7"}
    elif "CONTAINER" in normaliser(etat.modele) and "3M" in normaliser(etat.texte_ligne):
        feuille = "CONTAINER 3m 10pieds"  # container 3m -> onglet 3m (le 6m garde l'onglet par défaut)

    # --- En-tête ---
    # Chantier = intitulé + code postal + VILLE (la ville doit figurer dans la case chantier).
    chantier = " ".join(x for x in (entete.titre_chantier, entete.code_postal, entete.ville) if x)
    valeurs_entete = {
        "client": entete.client,
        "titre_chantier": chantier,
        "adresse": entete.adresse,
        "code_postal": entete.code_postal,
        "ville": entete.ville,
        "numero_offre": entete.numero_offre,
    }
    formats = cfg.get("entete_format") or {}
    for champ, cellule in entete_cellules.items():
        valeur = valeurs_entete.get(champ)
        if valeur and cellule:
            # Sur les vrais modèles, le libellé fait partie de la cellule (ex. « CLIENT : … ») :
            # un gabarit « CLIENT : {valeur} » reconstitue le libellé + la valeur.
            gabarit = formats.get(champ)
            a_ecrire[cellule] = gabarit.format(valeur=valeur) if gabarit else valeur

    # --- Titre du bloc (règle Joris) : réécrire la cellule-titre (ex. A6 des polysani) avec le
    #     contenu RÉEL lu sur la ligne du devis (« 5WC / 3UR »), à la place du texte figé du modèle. ---
    titre = cfg.get("titre_contenu") or {}
    if titre.get("cellule"):
        contenu = resumer_contenu_bloc(etat.texte_ligne)
        if contenu:
            gabarit = titre.get("gabarit") or "{contenu}"
            a_ecrire[titre["cellule"]] = gabarit.format(contenu=contenu)

    # --- Feuille chargée (lecture) pour les marquages repérés par TEXTE (positions variables) ---
    wb = load_workbook(modele_path, data_only=True)
    ws = wb[feuille] if feuille and feuille in wb.sheetnames else wb[wb.sheetnames[0]]

    # --- Fonction du bungalow ---
    cellule_fonction = cfg.get("fonction")
    if cellule_fonction and etat.fonction:
        # Modèle à ligne unique : on écrit la fonction retenue.
        a_ecrire[cellule_fonction] = etat.fonction
    elif etat.fonction:
        # R5 : garder la fonction retenue, EFFACER les autres cases de fonction (repérage par texte).
        for cell in regles_modeles.edits_fonction(ws, etat.fonction):
            a_ecrire[cell] = " "

    # --- Climatisation (R7) / Élingage point bas (R8) : « garder le bon, effacer l'autre » ---
    # Clim : on tranche TOUJOURS (mentionnée au devis -> OUI, sinon -> NON) — jamais les deux à la
    # fois. « inconnu/non applicable » (None, ex. sanitaire) = pas de clim -> NON. (edits_oui_non
    # ne fait rien si le modèle n'a pas de ligne CLIMATISE.)
    a_ecrire.update(regles_modeles.edits_oui_non(ws, "CLIMATISE", bool(etat.climatisation)))
    if etat.type_etat in ("individuel", "assemble"):
        # On répond toujours OUI/NON sur les modèles qui ont la ligne (sinon dict vide -> rien).
        a_ecrire.update(regles_modeles.edits_oui_non(ws, "ELINGAGE", etat.elingage_bas))

    # --- WC autonomes : mise en eau (= nb de WC), doses supp (R3), emplacements en trop (R2) ---
    if cfg.get("wc_emplacements"):
        # 1 mise en eau par WC de l'état (ex. 3 WC -> 3 mises en eau).
        a_ecrire.update(regles_modeles.edits_mise_en_eau(ws, etat.nb_modules))
    if etat.doses:
        a_ecrire.update(regles_modeles.edits_doses(ws, etat.doses))
    for cellule in (cfg.get("wc_emplacements") or [])[etat.nb_modules:]:
        a_ecrire[cellule] = " "  # R2 : effacer les emplacements au-delà du nombre de WC

    # --- Cuve : taille (3m/6m) écrite sur la ligne dédiée + RACCORDS = 1 si branchement au devis ---
    cuve = cfg.get("cuve") or {}
    if cuve:
        m = re.search(r"(\d+)\s*M", normaliser(etat.texte_ligne))
        if m and cuve.get("taille_cellule"):
            a_ecrire[cuve["taille_cellule"]] = f"CUVE {m.group(1)} M3"
        if entete.branchement and cuve.get("raccords_cellule"):
            cell = cuve["raccords_cellule"]
            label = str(ws[cell].value or "RACCORDS").strip()
            a_ecrire[cell] = f"{label} : 1"

    # --- Mobilier ---
    non_mappes: list[str] = []

    # 0) Kitchenette (réfectoire) : évier / chauffe-eau / micro-ondes / frigo = quantité du DEVIS
    #    (0 si absent), au lieu du « 1 » imprimé d'office. Ces articles sortent de la liste mobilier.
    motifs_kitchenette: list[str] = []
    for k in cfg.get("kitchenette") or []:
        motif = normaliser(k.get("motif") or "")
        qte = sum(m.quantite for m in etat.mobilier if motif and motif in normaliser(m.designation))
        a_ecrire[k["cellule"]] = f"{k['label']} : {qte}"
        if motif:
            motifs_kitchenette.append(motif)

    def _hors_kitchenette(mobilier):
        return [m for m in mobilier
                if not any(mo in normaliser(m.designation) for mo in motifs_kitchenette)]

    # 1) Zone d'inventaire des VRAIS modèles : la liste COMPLÈTE du devis est reportée telle
    #    quelle (« TABLE MODULAIRE RECT. 160X80 : 4 »), une ligne par cellule disponible ;
    #    s'il y a plus d'articles que de lignes, le reste est regroupé sur la dernière.
    zone = cfg.get("mobilier_zone") or []
    mobilier_liste = _hors_kitchenette(etat.mobilier) if motifs_kitchenette else etat.mobilier
    if zone and mobilier_liste:
        lignes = _lignes_inventaire(mobilier_liste)
        if len(zone) == 1:
            a_ecrire[zone[0]] = "\n".join(lignes)
        else:
            if len(lignes) > len(zone):
                lignes = lignes[: len(zone) - 1] + [" · ".join(lignes[len(zone) - 1 :])]
            for cellule, texte in zip(zone, lignes):
                a_ecrire[cellule] = texte

    # 2) Cases de quantité classiques (modèles qui en ont : cellule par mot-clé).
    table_mobilier = cfg.get("mobilier") or {}
    if table_mobilier:
        sommes: dict[str, int] = {}
        for item in etat.mobilier:
            cellule = _cellule_mobilier(item.designation, table_mobilier)
            if cellule:
                sommes[cellule] = sommes.get(cellule, 0) + item.quantite
            else:
                non_mappes.append(item.designation)
        a_ecrire.update(sommes)

    patch_xlsx.ecrire_cellules(
        modele_path, sortie_path, feuille, a_ecrire,
        hauteur_min_lignes=set(cfg.get("hauteur_lignes") or []),
    )
    return non_mappes
