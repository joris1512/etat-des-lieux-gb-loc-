"""Génère des MODÈLES FACTICES .xlsx dans templates/ pour faire tourner l'outil dès aujourd'hui.

⚠️ Ce sont des maquettes : lundi, on remplace ces fichiers par les VRAIS modèles GB
   (mêmes noms de fichiers) et on ajuste les cellules dans config/cellules.yaml.

Les libellés sont placés juste à gauche / au-dessus des cellules définies dans cellules.yaml,
pour que le remplissage soit visuellement cohérent dans la démo.
"""

from __future__ import annotations

import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

RACINE = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RACINE))

from app.config import CONFIG_CELLULES, CORRESPONDANCES_CSV, TEMPLATES_DIR  # noqa: E402
from app.correspondance import charger_correspondances  # noqa: E402

CHAMP_LIBELLE = {
    "client": "Client",
    "titre_chantier": "Chantier",
    "adresse": "Adresse",
    "code_postal": "Code postal",
    "ville": "Ville",
    "numero_offre": "N° offre",
}


def _cfg_modele(cfg: dict, modele: str) -> dict:
    defaut = cfg.get("defaut", {}) or {}
    surcharge = (cfg.get("modeles", {}) or {}).get(modele, {}) or {}
    return {
        "entete": {**(defaut.get("entete") or {}), **(surcharge.get("entete") or {})},
        "mobilier": surcharge["mobilier"] if "mobilier" in surcharge else (defaut.get("mobilier") or {}),
    }


def construire(modele: str, cfg: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Etat des lieux"
    titre = Font(bold=True, size=13)
    gras = Font(bold=True)
    bord = Border(*(Side(style="thin", color="BBBBBB"),) * 4)
    jaune = PatternFill("solid", fgColor="FFF3CD")

    ws["A1"] = f"ÉTAT DES LIEUX — {modele}"
    ws["A1"].font = titre
    ws["A1"].fill = PatternFill("solid", fgColor="EFEFEF")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22

    mc = _cfg_modele(cfg, modele)

    # En-tête : libellé en colonne A, valeur dans la cellule prévue.
    for champ, cellule in (mc.get("entete") or {}).items():
        r, c = coordinate_to_tuple(cellule)
        ws.cell(row=r, column=1, value=CHAMP_LIBELLE.get(champ, champ)).font = gras
        cell = ws[cellule]
        cell.fill = jaune
        cell.border = bord

    # Mobilier : libellé en colonne D, valeur dans la cellule (colonne E) prévue.
    mob = mc.get("mobilier") or {}
    if mob:
        r0 = min(coordinate_to_tuple(c)[0] for c in mob.values())
        ws.cell(row=r0 - 1, column=4, value="MOBILIER").font = gras
        ws.cell(row=r0 - 1, column=5, value="Qté").font = gras
        for mot, cellule in mob.items():
            r, c = coordinate_to_tuple(cellule)
            ws.cell(row=r, column=4, value=mot.title()).font = gras
            cell = ws[cellule]
            cell.fill = jaune
            cell.border = bord

    # Bloc « ne pas pré-remplir » pour matérialiser ce que l'outil NE touche pas.
    ws["A22"] = "État réel / Réserves / Signatures — NON pré-rempli (rempli à la main)"
    ws["A22"].font = Font(italic=True, color="888888")

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 26 if col in (1, 4) else 16
    return wb


def noms_modeles(cfg: dict) -> set[str]:
    modeles = {e.modele for e in charger_correspondances(CORRESPONDANCES_CSV)}
    modeles.add(cfg.get("modele_assemble", "bungalow_assemble.xlsx"))
    return modeles


def main() -> None:
    cfg = yaml.safe_load(CONFIG_CELLULES.read_text(encoding="utf-8"))
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    for modele in sorted(noms_modeles(cfg)):
        wb = construire(modele, cfg)
        wb.save(TEMPLATES_DIR / modele)
        print(f"  modèle factice : {modele}")
    print(f"OK — {TEMPLATES_DIR}")


if __name__ == "__main__":
    main()
