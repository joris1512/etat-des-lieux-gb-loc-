"""Mode chauffeur : constat en 2 temps (début/fin), cases par partie, 2 signatures, PDF, endpoints."""

import zipfile

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import terrain
from app.assemblage import construire_plan
from app.extraction import charger_fixture
from app.generation import generer
from app.main import app

PNG_MINIMAL = b"\x89PNG\r\n\x1a\n" + b"\x00" * 64
JPEG_MINIMAL = b"\xff\xd8\xff\xe0" + b"\x00" * 64


def _document_bungalow():
    rapport, job_dir = generer(None, utiliser_fixture=True)
    par_nom = {e.nom_fichier: e for e in construire_plan(charger_fixture()).etats}
    nom = next(n for n in rapport.fichiers if par_nom[n].type_etat == "individuel")
    return job_dir, nom


def test_parties_detectees_et_regroupees():
    job_dir, nom = _document_bungalow()
    parties = terrain.analyser_parties(job_dir / nom, None)
    libelles = [p["libelle"] for p in parties]
    assert len(parties) >= 6
    assert any("Sol" in x for x in libelles)
    assert any("Radiateur" in x for x in libelles)
    # Les continuations sont regroupées : plus de « (suite N) ».
    assert all("(suite" not in x for x in libelles)
    # Chaque partie référence au moins une ligne du formulaire.
    assert all(p["lignes"] for p in parties)


def test_liste_s_adapte_au_type_de_bloc():
    """Un sanitaire n'a pas les mêmes parties qu'un bungalow (liste lue dans le formulaire)."""
    from app.config import TEMPLATES_DIR

    bureau = [p["libelle"] for p in terrain.analyser_parties(TEMPLATES_DIR / "bungalow_vide.xlsx", None)]
    douche = [p["libelle"] for p in terrain.analyser_parties(TEMPLATES_DIR / "douches.xlsx", None)]
    assert bureau != douche
    assert any("Radiateur" in x for x in bureau)  # propre au bureau


def test_constat_ecrit_par_phase_dans_la_bonne_colonne():
    job_dir, nom = _document_bungalow()
    doc = job_dir / nom
    dossier = terrain.dossier_constat(job_dir, nom)
    parties = terrain.analyser_parties(doc, None)
    sol = next(p for p in parties if "Sol" in p["libelle"])
    # DÉBUT (colonne C) : Sol = bon état.
    terrain.enregistrer_constat(doc, None, "debut", {sol["cle"]: {"etat": "bon", "note": ""}}, dossier)
    assert load_workbook(doc).active[f"C{sol['lignes'][0]}"].value == "Bon état"
    # FIN (colonne F) : Sol = cassé + précision.
    terrain.enregistrer_constat(doc, None, "fin", {sol["cle"]: {"etat": "casse", "note": "Rayure 20 cm"}}, dossier)
    assert load_workbook(doc).active[f"F{sol['lignes'][0]}"].value == "Cassé — Rayure 20 cm"
    with zipfile.ZipFile(doc) as z:  # logo + perspectives préservés
        assert any("drawing" in n for n in z.namelist())
    recharge = terrain.charger_constat(doc, None, dossier)
    sol2 = next(p for p in recharge["parties"] if p["cle"] == sol["cle"])
    assert sol2["debut"]["etat"] == "bon" and sol2["fin"]["etat"] == "casse"


def test_deux_signatures_coexistent_dans_le_document():
    job_dir, nom = _document_bungalow()
    doc = job_dir / nom
    dossier = terrain.dossier_constat(job_dir, nom)
    ancres = terrain._ancres_signature(doc, None)
    assert ancres.get("debut") and ancres.get("fin") and ancres["debut"] != ancres["fin"]
    terrain.enregistrer_signature(dossier, PNG_MINIMAL, "M. Départ", document=doc, phase="debut", accord=True)
    terrain.enregistrer_signature(dossier, PNG_MINIMAL, "M. Retour", document=doc, phase="fin")
    assert (dossier / "signature-debut.png").exists() and (dossier / "signature-fin.png").exists()
    with zipfile.ZipFile(doc) as z:
        noms = z.namelist()
        assert "xl/media/signature_debut.png" in noms
        assert "xl/media/signature_fin.png" in noms  # les deux ne s'écrasent pas
    data = terrain.charger_constat(doc, None, dossier)
    assert data["debut"]["signe"] and data["fin"]["signe"]
    assert data["debut"]["signataire"] == "M. Départ" and data["fin"]["signataire"] == "M. Retour"


def test_phase_verrouillee_apres_signature_mais_pas_l_autre():
    job_dir, nom = _document_bungalow()
    doc = job_dir / nom
    dossier = terrain.dossier_constat(job_dir, nom)
    cle = terrain.analyser_parties(doc, None)[0]["cle"]
    terrain.enregistrer_signature(dossier, PNG_MINIMAL, "X", document=doc, phase="debut")
    with pytest.raises(terrain.ConstatSigne):  # début figé
        terrain.enregistrer_constat(doc, None, "debut", {cle: {"etat": "bon"}}, dossier)
    terrain.enregistrer_constat(doc, None, "fin", {cle: {"etat": "sale"}}, dossier)  # fin encore ouverte


def test_photos_validees_et_numerotees(tmp_path):
    d = tmp_path / "constat"
    d.mkdir()
    assert terrain.ajouter_photo(d, JPEG_MINIMAL) == "photo-01.jpg"
    assert terrain.ajouter_photo(d, PNG_MINIMAL) == "photo-02.png"
    with pytest.raises(ValueError):
        terrain.ajouter_photo(d, b"<script>pas une image</script>")


def test_signature_invalide_refusee_et_pdf(tmp_path):
    d = tmp_path / "constat"
    d.mkdir()
    with pytest.raises(ValueError):
        terrain.enregistrer_signature(d, b"pas un png", "X")
    terrain.enregistrer_signature(d, PNG_MINIMAL, "M. Client", phase="debut")
    assert (d / "signature-debut.png").exists()
    pdf = terrain.generer_pdf(d, titre="Bungalow 1", sous_titre="ACME · Chantier X", societe="GB")
    assert pdf.exists() and pdf.read_bytes()[:4] == b"%PDF"


def test_endpoints_constat(monkeypatch):
    import app.main as main_mod
    from app import generation

    monkeypatch.setattr(main_mod, "SORTIES_DIR", generation.SORTIES_DIR)
    job_dir, nom = _document_bungalow()
    tc = TestClient(app)
    base = f"/terrain/{job_dir.name}/{nom}"

    d = tc.get(base)
    assert d.status_code == 200 and len(d.json()["parties"]) >= 6
    cle = d.json()["parties"][0]["cle"]

    r = tc.post(base, json={"phase": "debut", "saisies": {cle: {"etat": "bon", "note": ""}}})
    assert r.status_code == 200
    p0 = next(p for p in r.json()["parties"] if p["cle"] == cle)
    assert p0["debut"]["etat"] == "bon"

    # Photo refusée si fausse image ; acceptée sinon, puis servie ; traversal refusé.
    assert tc.post(base + "/photo", files={"fichier": ("x.jpg", b"<html>", "image/jpeg")}).status_code == 400
    r = tc.post(base + "/photo", files={"fichier": ("x.jpg", JPEG_MINIMAL, "image/jpeg")})
    assert r.status_code == 200 and r.json()["photos"] == ["photo-01.jpg"]
    assert tc.get(f"/terrain-fichier/{job_dir.name}/{nom}/photo-01.jpg").status_code == 200
    assert tc.get(f"/terrain-fichier/{job_dir.name}/{nom}/..%5Cgb.db").status_code in (400, 404)

    # Signature (phase début) + PDF.
    import base64

    image = "data:image/png;base64," + base64.b64encode(PNG_MINIMAL).decode()
    r = tc.post(base + "/signature", json={"image": image, "signataire": "M. Client", "phase": "debut"})
    assert r.status_code == 200 and r.json()["debut"]["signe"] is True
    # Début figé : réenregistrer la phase début renvoie 409.
    assert tc.post(base, json={"phase": "debut", "saisies": {cle: {"etat": "sale"}}}).status_code == 409
    r = tc.post(base + "/pdf")
    assert r.status_code == 200 and r.json()["pdf"] == "constat.pdf"
    assert tc.get(f"/terrain-fichier/{job_dir.name}/{nom}/constat.pdf").status_code == 200

    # Partage : réservé au poste local (TestClient n'est pas local -> 403).
    assert tc.post(base + "/partager").status_code == 403
