"""Mobilier sur les états : libellés courts « norme GB », quantités regroupées, rien de perdu."""

from pathlib import Path

from openpyxl import load_workbook

from app.assemblage import construire_plan
from app.extraction import charger_fixture
from app.generation import generer
from app.models import MobilierItem
from app.remplissage import _lignes_inventaire


def _plan_et_fichiers():
    rapport, job_dir = generer(None, utiliser_fixture=True)
    par_nom = {e.nom_fichier: e for e in construire_plan(charger_fixture()).etats}
    return rapport, job_dir, par_nom


def _ws(job_dir, nom):
    wb = load_workbook(job_dir / nom)
    return wb[wb.sheetnames[0]]


def test_libelles_courts_et_regroupement():
    lignes = _lignes_inventaire(
        [
            MobilierItem(designation="TABLE MODULAIRE RECT. 160X80", quantite=4),
            MobilierItem(designation="TABLE PLIANTE 120", quantite=1),  # regroupée avec l'autre
            MobilierItem(designation="ARMOIRE BASSE PORTES BATTANTES 1000*800*380", quantite=1),
            MobilierItem(designation="ACCESSOIRE EVIER + CHAUFFE EAU", quantite=2),
            MobilierItem(designation="OBJET INCONNU SPECIAL", quantite=3),  # repli : tel quel
        ]
    )
    assert lignes[0] == "Tables : 5"
    assert "Armoires bureau basses : 1" in lignes
    assert "Evier + chauffe-eau : 2" in lignes
    assert "OBJET INCONNU SPECIAL : 3" in lignes


def test_vestiaire_norme_gb():
    rapport, job_dir, par_nom = _plan_et_fichiers()
    nom = next(
        n for n, e in par_nom.items()
        if e.bloc == "VESTIAIRE HOMMES" and n in rapport.fichiers
    )
    a46 = str(_ws(job_dir, nom)["A46"].value)
    assert "Armoires doubles : 10" in a46
    assert "Bancs : 2" in a46


def test_assemble_sans_mobilier_lignes_intactes():
    # Règle métier : l'assemblé ne porte JAMAIS le mobilier -> lignes d'inventaire vierges.
    rapport, job_dir, par_nom = _plan_et_fichiers()
    nom = next(
        n for n, e in par_nom.items()
        if e.type_etat == "assemble" and e.bloc == "SALLE DE REUNION" and n in rapport.fichiers
    )
    ws = _ws(job_dir, nom)
    lignes = [str(ws[f"A{r}"].value).strip() for r in range(43, 48)]
    assert "Chaise :" in lignes  # libellés d'origine du modèle, non remplis
    assert not any(x.split(":")[-1].strip().isdigit() for x in lignes if ":" in x)


def test_salle_reunion_mobilier_sur_individuel():
    rapport, job_dir, par_nom = _plan_et_fichiers()
    etat = next(
        e for e in par_nom.values()
        if e.type_etat == "individuel" and e.bloc == "SALLE DE REUNION" and e.mobilier
    )
    a46 = str(_ws(job_dir, etat.nom_fichier)["A46"].value)
    assert "Tables : 4" in a46
    assert "Chaises : 20" in a46
    assert "Armoires bureau basses : 1" in a46


def test_refectoire_mobilier_sur_individuel_norme_gb():
    rapport, job_dir, par_nom = _plan_et_fichiers()
    etat = next(
        e for e in par_nom.values()
        if e.type_etat == "individuel" and e.bloc == "REFECTOIRE" and e.mobilier
    )
    assert etat.modele == "bungalow_refectoire_kitchenette.xlsx"
    ws = _ws(job_dir, etat.nom_fichier)
    contenu = str(ws["A49"].value) + " · " + str(ws["A50"].value)
    for attendu in ("Tables : 4", "Bancs : 8", "Frigo : 2", "Micro-ondes : 2", "Evier + chauffe-eau : 2"):
        assert attendu in contenu


def test_aucun_avertissement_mobilier_sur_eiffage():
    rapport, _job_dir, _ = _plan_et_fichiers()
    assert not any("rattach" in a.lower() for a in rapport.avertissements)


def _hauteur_zone_mobilier(ws):
    """Hauteur totale de la zone mobilier fusionnée A46:B47 (pt)."""
    return (ws.row_dimensions[46].height or 15) + (ws.row_dimensions[47].height or 15)


def test_mobilier_abondant_renvoie_a_la_ligne(tmp_path):
    """Beaucoup de mobilier = renvoi à la ligne + ZONE agrandie (pas « à la suite »), sans casser
    le classeur. La case A46 est fusionnée A46:B47 : on vérifie la hauteur TOTALE de la zone."""
    import re
    import zipfile

    from app import patch_xlsx

    src = Path(__file__).parent.parent / "templates" / "bungalow_mobilier.xlsx"
    out = tmp_path / "mobilier.xlsx"
    lignes = [f"Article {i} : {i}" for i in range(1, 13)]  # 12 lignes
    patch_xlsx.ecrire_cellules(src, out, "Bungalow avec mobiliers", {"A46": "\n".join(lignes)})

    wb = load_workbook(out)
    ws = wb["Bungalow avec mobiliers"]
    assert ws["A46"].alignment.wrap_text is True          # renvoi à la ligne activé
    assert str(ws["A46"].value).count("\n") == 11          # 12 lignes conservées
    assert _hauteur_zone_mobilier(ws) >= 12 * 15           # la ZONE fusionnée est assez haute

    # Intégrité : le logo / les dessins du modèle restent présents.
    with zipfile.ZipFile(out) as z:
        assert any(re.search(r"drawing\d*\.xml$", n) for n in z.namelist())
        assert any(n.startswith("xl/media/") for n in z.namelist())


def test_mobilier_court_ne_pousse_pas_le_contenu_du_dessous(tmp_path):
    """Peu de mobilier (tient dans la zone déjà prévue) : la hauteur n'est PAS augmentée →
    la section « matériel à assurer » et les signatures en dessous ne bougent pas."""
    from app import patch_xlsx

    src = Path(__file__).parent.parent / "templates" / "bungalow_mobilier.xlsx"
    ref = load_workbook(src)["Bungalow avec mobiliers"]
    zone_avant = (ref.row_dimensions[46].height or 15) + (ref.row_dimensions[47].height or 15)

    out = tmp_path / "mobilier.xlsx"
    patch_xlsx.ecrire_cellules(src, out, "Bungalow avec mobiliers",
                               {"A46": "Armoires doubles : 10\nBancs : 2"})  # 2 lignes
    ws = load_workbook(out)["Bungalow avec mobiliers"]
    assert ws["A46"].alignment.wrap_text is True
    assert _hauteur_zone_mobilier(ws) == zone_avant  # zone INCHANGÉE : rien poussé en dessous


def test_kit_reste_choisi_quand_kitchenette_seule():
    from app.assemblage import _kitchenette_seulement

    frigo = [MobilierItem(designation="ACCESSOIRE REFRIGERATEUR", quantite=1)]
    mixte = frigo + [MobilierItem(designation="TABLE MODULAIRE", quantite=2)]
    assert _kitchenette_seulement(frigo) is True
    assert _kitchenette_seulement(mixte) is False
    assert _kitchenette_seulement([]) is False
