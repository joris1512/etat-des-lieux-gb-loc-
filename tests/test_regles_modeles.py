"""Marquage sur les VRAIS modèles : garder la bonne option, effacer les autres (R5/R7/R8/R1)."""

import glob
from pathlib import Path

from openpyxl import load_workbook

from app.regles_modeles import edits_fonction, edits_oui_non, onglet_douches

MR = Path(__file__).parent.parent / "modeles_reels"


def _ws(motif):
    fichiers = glob.glob(str(MR / motif))
    assert fichiers, f"modèle introuvable : {motif}"
    wb = load_workbook(fichiers[0], data_only=True)
    return wb[wb.sheetnames[0]]


def test_r5_efface_les_autres_fonctions_bungalow_vide():
    ws = _ws("bungalow/*BUNG VIDE.xlsx")  # B8=SALLE DE REUNION D8=VESTIAIRE E8=BUREAU G8=REFECTOIRE
    edits = edits_fonction(ws, "BUREAU")
    assert set(edits) == {"B8", "D8", "G8"}  # on efface les 3 autres, on garde BUREAU (E8)
    assert all(v is None for v in edits.values())


def test_r5_ordre_different_8m2():
    ws = _ws("bungalow/*8M*VIDE*")  # B8=REFECTOIRE D8=BUREAU E8=SALLE DE REUNION G8=VESTIAIRE
    edits = edits_fonction(ws, "VESTIAIRE")
    assert set(edits) == {"B8", "D8", "E8"}  # garde VESTIAIRE (G8)


def test_r5_fonction_absente_ne_touche_rien():
    ws = _ws("bungalow/*BUNG VIDE.xlsx")
    assert edits_fonction(ws, None) == {}
    assert edits_fonction(ws, "FONCTION INEXISTANTE") == {}  # aucune case ne correspond -> tout effacé ?


def test_r7_clim_oui_efface_non():
    ws = _ws("bungalow/*BUNG VIDE.xlsx")
    edits = edits_oui_non(ws, "CLIMATISE", oui=True)
    assert list(edits) == ["A9"]
    val = edits["A9"].upper()
    assert "OUI" in val and "NON" not in val


def test_r7_clim_non_efface_oui():
    ws = _ws("bungalow/*BUNG VIDE.xlsx")
    edits = edits_oui_non(ws, "CLIMATISE", oui=False)
    val = edits["A9"].upper()
    assert "NON" in val and "OUI" not in val


def test_r8_elingage_present_dans_vide():
    ws = _ws("bungalow/*BUNG VIDE.xlsx")  # A7 = ELINGAGE PT BAS : OUI NON
    edits = edits_oui_non(ws, "ELINGAGE", oui=True)
    assert list(edits) == ["A7"]
    assert "NON" not in edits["A7"].upper()


def test_r8_elingage_absent_du_8m2():
    ws = _ws("bungalow/*8M*VIDE*")  # pas de ligne élingage
    assert edits_oui_non(ws, "ELINGAGE", oui=True) == {}


def test_r1_onglet_douches():
    assert onglet_douches(4) == "Petit Sanitaire Douche"
    assert onglet_douches(6) == "Grand Sanitaire Douche"
