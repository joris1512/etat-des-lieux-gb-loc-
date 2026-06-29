"""Test de bout en bout sur le devis EIFFAGE (via la fixture) — critères d'acceptation v1."""

from openpyxl import load_workbook

from app.assemblage import construire_plan
from app.extraction import charger_fixture
from app.generation import generer
from app.remplissage import _cfg_modele


def _ouvrir_etat(job_dir, nom, etat):
    """Ouvre le fichier produit sur le bon onglet (selon la config du modèle)."""
    cfg = _cfg_modele(etat.modele)
    wb = load_workbook(job_dir / nom)
    ws = wb[cfg["feuille"]] if cfg.get("feuille") else wb.active
    return ws, cfg


def test_plan_eiffage_compte_et_types():
    extraction = charger_fixture()
    plan = construire_plan(extraction)

    # 11 bungalows standard -> 11 individuels ; 4 blocs assemblés ; 2 sanitaires = 17 fichiers.
    types = [e.type_etat for e in plan.etats]
    assert types.count("individuel") == 11
    assert types.count("assemble") == 4
    assert types.count("sanitaire") == 2
    assert len(plan.etats) == 17
    assert plan.non_reconnus == []


def test_assembles_attendus():
    plan = construire_plan(charger_fixture())
    assembles = {e.bloc for e in plan.etats if e.type_etat == "assemble"}
    assert assembles == {"REFECTOIRE", "2 BUREAUX INDEPENDANTS", "2 BUREAUX", "SALLE DE REUNION"}


def test_sanitaires_bons_modeles():
    plan = construire_plan(charger_fixture())
    modeles = sorted(e.modele for e in plan.etats if e.type_etat == "sanitaire")
    assert modeles == ["sanitaire_2wc_2d_2u.xlsx", "sanitaire_2wc_pmr.xlsx"]


def test_salle_reunion_assemble_porte_20_chaises():
    plan = construire_plan(charger_fixture())
    sr = next(e for e in plan.etats if e.type_etat == "assemble" and e.bloc == "SALLE DE REUNION")
    chaises = sum(m.quantite for m in sr.mobilier if "CHAISE" in m.designation.upper())
    assert chaises == 20


def test_generation_complete_produit_17_fichiers_et_zip():
    rapport, job_dir = generer(None, utiliser_fixture=True)
    assert len(rapport.fichiers) == 17
    assert rapport.zip_nom is not None
    assert (job_dir / rapport.zip_nom).exists()

    par_nom = {e.nom_fichier: e for e in construire_plan(charger_fixture()).etats}
    # En-tête écrite sur un bungalow réel : cellule client E5, préfixée « Client : ».
    nom_bung = next(n for n in rapport.fichiers if par_nom[n].type_etat == "individuel")
    ws, cfg = _ouvrir_etat(job_dir, nom_bung, par_nom[nom_bung])
    assert ws[cfg["entete"]["client"]].value == "Client : EIFFAGE TRX MARITIMES FLUVIAUX"


def test_corps_inspection_non_rempli():
    rapport, job_dir = generer(None, utiliser_fixture=True)
    par_nom = {e.nom_fichier: e for e in construire_plan(charger_fixture()).etats}
    nom_bung = next(n for n in rapport.fichiers if par_nom[n].type_etat == "individuel")
    ws, _ = _ouvrir_etat(job_dir, nom_bung, par_nom[nom_bung])
    # La grille d'inspection (corps) n'est jamais pré-remplie : libellé d'origine intact.
    assert ws["A40"].value == "Radiateur"


def test_fonction_bungalow_ecrite_dans_les_fichiers_produits():
    rapport, job_dir = generer(None, utiliser_fixture=True)
    # Plan déterministe : on relie chaque fichier produit à son état pour connaître la fonction.
    par_nom = {e.nom_fichier: e for e in construire_plan(charger_fixture()).etats}

    verifies = 0
    for nom in rapport.fichiers:
        etat = par_nom[nom]
        if not etat.fonction:
            continue
        ws, cfg = _ouvrir_etat(job_dir, nom, etat)
        # La fonction retenue REMPLACE la ligne d'origine (mot seul, ex. « VESTIAIRE »).
        assert ws[cfg["fonction"]].value == etat.fonction
        verifies += 1
    assert verifies >= 4  # au moins les 4 blocs assemblés (réfectoire, 2x bureaux, réunion)
